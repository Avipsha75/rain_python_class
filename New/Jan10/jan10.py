import openpyxl

#Create a messy dataset in Excel for demonstration
def create_messy_dataset(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales Data"

    #Add messy data
    sheet.append(["Date", "Product", "Quantity", "Price"])
    sheet.append(["2025-01-05", "Laptop", 3, 1200.50])
    sheet.append(["", "Mouse", "", 25.99]) #Missing date and quantity
    sheet.append(["2025-01-05", "Laptop", 3, 1200.50]) # Duplicate row
    sheet.append(["2025-01-06", "Keyboard", 2, None]) # Missing Price

    workbook.save(file_path)
    print("Messy dataset created successfully.")

# Clean the dataset
def clean_dataset(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    cleaned_data  = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #Skip rows with missing data
        if None in row or "" in row:
            continue
        #Remove duplicates
        if row not in cleaned_data:
            cleaned_data.append(row)

    #Clear the sheet and write cleaned data
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    
    #Re-add headers
    headers = ["Date", "Product", "Quantity", "Price"]
    sheet.append(headers)

    #Add cleaned data
    for row in cleaned_data:
        sheet.append(row)

    workbook.save(file_path)
    print("Dataset cleaned.")

#Example usuage
file_path = "messy_sales_data.xlsx"
create_messy_dataset(file_path)
clean_dataset(file_path)